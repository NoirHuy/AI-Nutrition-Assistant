"""
excel_to_txt.py
Chuyển đổi bảng thành phần dinh dưỡng (Excel) → văn bản mô tả (.txt)
để làm đầu vào cho pipeline trích xuất Knowledge Graph (EDC Framework).

Cách dùng:
    python edc-main/excel_to_txt.py \
        --input docs/food_vietnam_final.xlsx \
        --output edc-main/datasets/food_vietnam.txt \
        --chunk_size 20
"""

import argparse
import openpyxl
import os

# Ánh xạ: tên cột → (tên hiển thị, ngưỡng cao, nhận xét cao, nhận xét thấp)
NUTRIENT_MAP = {
    "Calories (kcal)":       ("năng lượng",       300,  "giàu năng lượng",       "ít năng lượng"),
    "Protein (g)":           ("protein",            10,  "giàu protein",           "ít protein"),
    "Fat (g)":               ("chất béo",           10,  "nhiều chất béo",         "ít chất béo"),
    "Carbonhydrates (g)":    ("carbohydrate",       30,  "giàu carbohydrate",      "ít carbohydrate"),
    "Chất xơ (g)":           ("chất xơ",             3,  "giàu chất xơ",           "ít chất xơ"),
    "Cholesterol (mg)":      ("cholesterol",        50,  "nhiều cholesterol",      "ít cholesterol"),
    "Canxi (mg)":            ("canxi",             100,  "giàu canxi",             "ít canxi"),
    "Photpho (mg)":          ("phospho",           100,  "giàu phospho",           "ít phospho"),
    "Sắt (mg)":              ("sắt",                 3,  "giàu sắt",               "ít sắt"),
    "Natri (mg)":            ("natri",             200,  "nhiều natri",            "ít natri"),
    "Kali (mg)":             ("kali",              200,  "giàu kali",              "ít kali"),
    "Beta Caroten (mcg)":    ("beta caroten",      100,  "giàu beta caroten",      "ít beta caroten"),
    "Vitamin A (mcg)":       ("vitamin A",          50,  "giàu vitamin A",         "ít vitamin A"),
    "Vitamin B1 (mg)":       ("vitamin B1",        0.1,  "giàu vitamin B1",        "ít vitamin B1"),
    "Vitamin C (mg)":        ("vitamin C",          10,  "giàu vitamin C",         "ít vitamin C"),
}


def safe_float(val) -> float:
    """Chuyển giá trị ô excel (có thể là str như '8,6') sang float."""
    if val is None:
        return 0.0
    try:
        return float(str(val).replace(",", "."))
    except ValueError:
        return 0.0


def food_to_paragraph(row: dict) -> str:
    """Chuyển 1 hàng Excel → 1 đoạn văn mô tả."""
    name = str(row.get("TÊN THỨC ĂN", "")).strip()
    category = str(row.get("Loại", "")).strip()
    if not name:
        return ""

    lines = []

    # Câu mở đầu
    if category and category != "None":
        lines.append(f"{name} là một thực phẩm thuộc nhóm {category}.")
    else:
        lines.append(f"{name} là một thực phẩm phổ biến trong ẩm thực Việt Nam.")

    # Mô tả từng chất dinh dưỡng
    nutrients_high = []
    nutrients_detail = []

    for col, (viet_name, threshold, label_high, label_low) in NUTRIENT_MAP.items():
        val = safe_float(row.get(col))
        unit_str = col.split("(")[-1].replace(")", "").strip() if "(" in col else ""
        val_str = str(row.get(col, "")).strip().replace(",", ".")

        if val > 0:
            nutrients_detail.append(f"{viet_name}: {val_str} {unit_str}")
            if val >= threshold:
                nutrients_high.append(viet_name)

    # Câu thành phần chi tiết
    if nutrients_detail:
        lines.append(f"Thành phần dinh dưỡng trong 100g {name} bao gồm: {', '.join(nutrients_detail)}.")

    # Câu nhận xét đặc trưng
    if nutrients_high:
        lines.append(f"{name} được xem là nguồn {', '.join(nutrients_high)} tốt.")

    # Câu về bệnh lý liên quan (dựa trên đặc trưng)
    nat = safe_float(row.get("Natri (mg)"))
    fat = safe_float(row.get("Fat (g)"))
    sugar = safe_float(row.get("Carbonhydrates (g)"))
    fiber = safe_float(row.get("Chất xơ (g)"))
    chol = safe_float(row.get("Cholesterol (mg)"))

    if nat > 400:
        lines.append(f"Do chứa nhiều natri ({str(row.get('Natri (mg)', '')).strip()} mg), "
                     f"{name} có thể làm trầm trọng thêm bệnh tăng huyết áp và bệnh thận.")
    if fat > 15 or chol > 100:
        lines.append(f"{name} chứa nhiều chất béo và cholesterol, "
                     f"cần hạn chế ở người mắc bệnh tim mạch và gan nhiễm mỡ.")
    if sugar > 50:
        lines.append(f"Hàm lượng carbohydrate cao ({str(row.get('Carbonhydrates (g)', '')).strip()} g) "
                     f"trong {name} có thể ảnh hưởng đến đường huyết, "
                     f"người bệnh tiểu đường cần thận trọng khi sử dụng.")
    if fiber >= 3:
        lines.append(f"Chất xơ trong {name} ({str(row.get('Chất xơ (g)', '')).strip()} g) "
                     f"có tác dụng hỗ trợ kiểm soát đường huyết và phòng ngừa táo bón.")

    return " ".join(lines)


def convert(input_path: str, output_path: str, chunk_size: int = 20):
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    # Đọc header
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    print(f"📋 Cột: {headers}")
    print(f"📊 Tổng số hàng dữ liệu: {ws.max_row - 1}")

    paragraphs = []
    skipped = 0

    for row_idx in range(2, ws.max_row + 1):
        row = {headers[c]: ws.cell(row_idx, c + 1).value for c in range(len(headers))}
        para = food_to_paragraph(row)
        if para:
            paragraphs.append(para)
        else:
            skipped += 1

    # Ghi ra file txt, mỗi chunk_size món là 1 đoạn lớn (để phù hợp với pipeline EDC)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        for i in range(0, len(paragraphs), chunk_size):
            chunk = paragraphs[i:i + chunk_size]
            f.write("\n".join(chunk))
            f.write("\n\n")  # Blank line phân tách chunk

    print(f"✅ Đã ghi {len(paragraphs)} món ăn → {output_path}")
    print(f"⚠️  Bỏ qua {skipped} hàng rỗng")
    print(f"📦 Tổng chunks (~{chunk_size} món/chunk): {(len(paragraphs) + chunk_size - 1) // chunk_size}")

    # Preview 2 đoạn đầu
    print("\n--- PREVIEW ---")
    for p in paragraphs[:2]:
        print(p)
        print()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Chuyển Excel dinh dưỡng → TXT cho EDC pipeline")
    parser.add_argument("--input",       default="docs/food_vietnam_final.xlsx")
    parser.add_argument("--output",      default="edc-main/datasets/food_vietnam.txt")
    parser.add_argument("--chunk_size",  type=int, default=20,
                        help="Số món ăn mỗi đoạn (để chia cho EDC pipeline)")
    args = parser.parse_args()
    convert(args.input, args.output, args.chunk_size)
